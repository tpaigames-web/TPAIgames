from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colors ──
HEADER_FILL = PatternFill('solid', fgColor='2F4F2F')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
DATA_FONT = Font(name='Arial', size=10)
BLUE_FONT = Font(name='Arial', size=10, color='0000FF')
RARITY_FILLS = {
    '白': PatternFill('solid', fgColor='F0F0F0'),
    '绿': PatternFill('solid', fgColor='E8F5E9'),
    '蓝': PatternFill('solid', fgColor='E3F2FD'),
    '紫': PatternFill('solid', fgColor='F3E5F5'),
    '橙': PatternFill('solid', fgColor='FFF3E0'),
}
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data(ws, row, max_col, fill=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if fill:
            cell.fill = fill

# ── Tower Data ──
TOWERS = [
    {'name': '稻草人', 'emoji': '🌾', 'rarity': '白', 'cost': 100, 'dmg': 15, 'interval': 0.8, 'range': 180, 'atk_type': '地面', 'hero': '否', 'path_only': '否', 'unlock': 1,
     'paths': [
         {'name': '攻击提升', 'dmg_b': 0.12, 'spd_b': 0, 'rng_b': 0, 'costs': [100,200,350,550,800],
          'tiers': ['精准投掷','双倍草束','稳定摆臂','锋利草刃','终极收割'],
          'effects': ['伤害+12%','伤害累计+12%','伤害累计+12%','伤害累计+12%','伤害累计+12%'], 'air_change': None},
         {'name': '攻击速度', 'dmg_b': 0, 'spd_b': 0.10, 'rng_b': 0, 'costs': [80,160,280,450,650],
          'tiers': ['快速摆动','节奏加快','稻草旋风','极速摇摆','永动稻草人'],
          'effects': ['攻速+10%','攻速累计+10%','攻速累计+10%','攻速累计+10%','攻速累计+10%'], 'air_change': None},
         {'name': '射程提升', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [90,180,300,500,750],
          'tiers': ['手臂延伸','长臂草人','远程覆盖','广域守望','全场守护'],
          'effects': ['射程+15%','射程累计+15%','射程累计+15%','射程累计+15%','射程累计+15%'], 'air_change': None},
         {'name': '增加效果', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['缓步枯草','强化缠绕','巨物捆绑','防空稻草','全域减速'],
          'effects': ['减速10%','已减速目标20%','可对大型减速','可攻击空中','全类型减速'], 'air_change': 4},
     ]},
    {'name': '蜂巢', 'emoji': '🍯', 'rarity': '白', 'cost': 120, 'dmg': 12, 'interval': 1.5, 'range': 200, 'atk_type': '全部', 'hero': '否', 'path_only': '否', 'unlock': 1,
     'paths': [
         {'name': '蜂群强化', 'dmg_b': 0.12, 'spd_b': 0, 'rng_b': 0, 'costs': [100,200,350,550,800],
          'tiers': ['大型蜂群','双倍蜂巢','毒针增强','蜂王护卫','蜂巢领域'],
          'effects': ['蜜蜂+1,伤害提升','每只伤害+15%','中毒3dps/3s','召唤蜂王2.5x伤害','蜂群极限'], 'air_change': None},
         {'name': '攻击速度', 'dmg_b': 0, 'spd_b': 0.12, 'rng_b': 0, 'costs': [80,160,280,450,650],
          'tiers': ['快速蜂群','多路出击','蜂群涌出','蜂海战术','无尽蜂涌'],
          'effects': ['攻速+15%','攻速大幅提升','攻速再提升','持续攻击','攻速极限'], 'air_change': None},
         {'name': '飞行射程', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [90,180,300,480,700],
          'tiers': ['飞行距离','远程追踪','蜂群领地','广域蜂巢','全场蜂群'],
          'effects': ['射程+15%','巡逻更远','追踪大幅提升','全方位覆盖','射程极大'], 'air_change': None},
         {'name': '毒素增强', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [120,250,420,650,950],
          'tiers': ['毒针蜜蜂','强效毒素','叠加中毒','神经毒素','致命毒巢'],
          'effects': ['蜂毒3dps/4s','5dps/5s','可叠2层','中毒减速20%','叠3层+毒爆'], 'air_change': None},
     ]},
    {'name': '水压管道', 'emoji': '🚿', 'rarity': '绿', 'cost': 150, 'dmg': 0, 'interval': 1.5, 'range': 200, 'atk_type': '全部', 'hero': '否', 'path_only': '否', 'unlock': 1,
     'paths': [
         {'name': '高压水枪', 'dmg_b': 0.20, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['高压水枪','多向高压','高压水炮','液氮喷射','超高压洪流'],
          'effects': ['单体少量伤害','3目标伤害减速','小范围击退减速','减速40%','大范围强力击退'], 'air_change': None},
         {'name': '速度', 'dmg_b': 0, 'spd_b': 0.10, 'rng_b': 0, 'costs': [100,200,350,550,800],
          'tiers': ['加压泵','高效水泵','涡轮增压','极速喷射','无限水流'],
          'effects': ['喷射+10%','累计+10%','5目标','6目标','8目标'], 'air_change': None},
         {'name': '射程', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [90,180,300,480,700],
          'tiers': ['延长水管','加压延伸','远程水枪','广域喷射','全场覆盖'],
          'effects': ['射程+15%','累计+15%','减速+1.5s','减速+3s','减速+4.5s'], 'air_change': None},
         {'name': '效果增益', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [120,250,400,600,900],
          'tiers': ['强化减速','冰冷之水','冰封之力','群体蔓延','永恒寒冰'],
          'effects': ['减速30%','减速40%/3s','50%冻结5s','扩散100px','50%减速5s/150px'], 'air_change': None},
     ]},
    {'name': '带刺铁网', 'emoji': '🪤', 'rarity': '绿', 'cost': 150, 'dmg': 5, 'interval': 0.5, 'range': 120, 'atk_type': '地面', 'hero': '否', 'path_only': '是', 'unlock': 5,
     'paths': [
         {'name': '刺伤强化', 'dmg_b': 0.12, 'spd_b': 0, 'rng_b': 0, 'costs': [120,240,400,600,900],
          'tiers': ['锋利铁刺','穿透铁刺','出血铁网','深度出血','终极铁刺'],
          'effects': ['伤害+15%','无视10%护甲','出血2dps','4dps叠2层','3层叠加灼伤'], 'air_change': None},
         {'name': '减速铁网', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [100,200,350,550,800],
          'tiers': ['缠绕减速','强力缠绕','持续拖拽','重度减速','极致缠绕'],
          'effects': ['减速20%','减速30%','残留+1s','被减速+15%伤害','减速50%'], 'air_change': None},
         {'name': '铺设范围', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [90,180,300,480,700],
          'tiers': ['扩展铁网','大型铁网','广域铺设','超级铁网','全场铁幕'],
          'effects': ['范围+15%','扩大','大幅提升','封锁通道','多路封堵'], 'air_change': None},
         {'name': '强化铁网', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['自动修复','爆炸铁网','带电铁丝','双层铁网','磁化铁网'],
          'effects': ['3s自动修复','小范围爆炸','30%麻痹0.5s','伤害翻倍+减速','拉近敌人'], 'air_change': None},
     ]},
    {'name': '农场大炮', 'emoji': '💣', 'rarity': '绿', 'cost': 250, 'dmg': 35, 'interval': 0.6, 'range': 240, 'atk_type': '地面', 'hero': '否', 'path_only': '否', 'unlock': 10,
     'paths': [
         {'name': '炮弹威力', 'dmg_b': 0.15, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['重型炮弹','穿甲弹头','燃烧炮弹','超级穿甲','末日炮击'],
          'effects': ['伤害+15%','无视10%护甲','燃烧3s','无视30%护甲','击杀爆炸'], 'air_change': None},
         {'name': '射程提升', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [120,240,400,600,900],
          'tiers': ['加长炮管','精准弹道','远程炮击','超远程炮','洲际导弹'],
          'effects': ['射程+15%','累计提升','大幅增加','大半地图','全图覆盖'], 'air_change': None},
         {'name': '榴弹炮', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [180,360,600,850,1200],
          'tiers': ['爆炸弹头','溅射扩大','燃烧弹幕','磷弹攻击','核心炮击'],
          'effects': ['小范围溅射','溅射扩大','燃烧3s','持续伤害大幅提升','大爆炸+高燃烧'], 'air_change': None},
         {'name': '射速提升', 'dmg_b': 0, 'spd_b': 0.15, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['自动装填','快速弹药','连续炮击','速射模式','极速炮兵'],
          'effects': ['射速+15%','累计提升','大幅提升','不停歇','输出极限'], 'air_change': None},
     ]},
    {'name': '捕兽夹', 'emoji': '🪤', 'rarity': '绿', 'cost': 200, 'dmg': 350, 'interval': 10.0, 'range': 60, 'atk_type': '地面', 'hero': '否', 'path_only': '是', 'unlock': 5,
     'paths': [
         {'name': '高级捕兽夹', 'dmg_b': 0.10, 'spd_b': 0, 'rng_b': 0, 'costs': [200,400,650,900,1300],
          'tiers': ['高级捕兽夹','钢铁陷阱','爆炸夹','毒液夹','终极钢夹'],
          'effects': ['范围AOE','范围扩大+25%','爆炸+50%','中毒持续伤害','极大范围毁灭'], 'air_change': None},
         {'name': '重置速度', 'dmg_b': 0, 'spd_b': 0.80, 'rng_b': 0, 'costs': [120,250,400,600,900],
          'tiers': ['快速复位','弹簧加速','自动复位','极速复位','永不停歇'],
          'effects': ['冷却~6s','~4s','~3s','~2.5s','2s'], 'air_change': None},
         {'name': '伤害提升', 'dmg_b': 0.20, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['锋利钢齿','内置尖刺','爆炸装置','毒液涂层','终极钢齿'],
          'effects': ['伤害+25%','伤害+50%出血','伤害+75%爆炸','伤害+100%中毒','大型必杀'], 'air_change': None},
         {'name': '毒刺夹', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [200,400,650,900,1300],
          'tiers': ['毒刺附着','强效毒素','毒气扩散','致命毒素','毒液炸弹'],
          'effects': ['毒20dps/4s','35dps/5s','80px范围中毒','减速30%','120px毒爆+眩晕5s'], 'air_change': None},
     ]},
    {'name': '播种机', 'emoji': '🌱', 'rarity': '蓝', 'cost': 350, 'dmg': 15, 'interval': 0.9, 'range': 350, 'atk_type': '地面', 'hero': '否', 'path_only': '否', 'unlock': 15,
     'paths': [
         {'name': '种子威力', 'dmg_b': 0.15, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['重型种子','爆裂种子','溅射扩大','标记弹头','终极炮射'],
          'effects': ['伤害+15%','小范围溅射','爆裂扩大','标记+15%增伤','极大范围+标记'], 'air_change': None},
         {'name': '播射速度', 'dmg_b': 0, 'spd_b': 0.12, 'rng_b': 0, 'costs': [120,240,400,600,900],
          'tiers': ['快速播种','高速输送','连续播射','全自动化','超速播种机'],
          'effects': ['攻速+12%','累计提升','几乎无间隔','大幅提升','攻速极限'], 'air_change': None},
         {'name': '射程提升', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [100,200,350,550,800],
          'tiers': ['加长枪管','精准弹道','远程播种','超远程炮','全图覆盖'],
          'effects': ['射程+15%','累计提升','更大区域','大幅提升','无死角'], 'air_change': None},
         {'name': '发芽效果', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [180,360,600,850,1200],
          'tiers': ['生根发芽','荆棘草丛','减速藤蔓','爆炸草丛','超级草丛'],
          'effects': ['草丛阻碍','3dps伤害','移速-30%','3s后爆炸','全效果叠加'], 'air_change': None},
     ]},
    {'name': '蘑菇炸弹', 'emoji': '🍄', 'rarity': '蓝', 'cost': 350, 'dmg': 25, 'interval': 0.4, 'range': 220, 'atk_type': '地面', 'hero': '否', 'path_only': '否', 'unlock': 25,
     'paths': [
         {'name': '爆炸威力', 'dmg_b': 0.15, 'spd_b': 0, 'rng_b': 0, 'costs': [180,360,600,850,1200],
          'tiers': ['强力炸药','爆炸增幅','碎片飞溅','连锁爆炸','末日蘑菇'],
          'effects': ['伤害+15%','累计提升','二次击中','击杀引爆','伤害+范围极限'], 'air_change': None},
         {'name': '投掷速度', 'dmg_b': 0, 'spd_b': 0.12, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['快速投掷','连续投掷','自动投弹','高速投掷机','极速轰炸'],
          'effects': ['攻速+12%','累计提升','减少装填','大幅提升','如雨而降'], 'air_change': None},
         {'name': '爆炸范围', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [160,320,540,780,1100],
          'tiers': ['扩大爆心','大型炸弹','广域爆破','超级炸弹','核爆蘑菇'],
          'effects': ['半径+15%','累计扩大','大范围','覆盖区域','无处可逃'], 'air_change': None},
         {'name': '孢子毒气', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [200,400,650,900,1300],
          'tiers': ['毒气云','浓缩毒素','扩散毒雾','减速毒气','剧毒孢子'],
          'effects': ['毒3dps/4s','5dps','范围+50%','移速-25%','10s大范围+眩晕'], 'air_change': None},
     ]},
    {'name': '风车', 'emoji': '🌀', 'rarity': '蓝', 'cost': 350, 'dmg': 12, 'interval': 1.3, 'range': 250, 'atk_type': '全部', 'hero': '否', 'path_only': '否', 'unlock': 20,
     'paths': [
         {'name': '旋风伤害', 'dmg_b': 0.12, 'spd_b': 0, 'rng_b': 0, 'costs': [150,300,500,700,1000],
          'tiers': ['锋利刀刃','双层风刃','穿透旋风','狂风刀阵','终极旋风'],
          'effects': ['伤害+12%','累计提升','穿透多敌','大幅提升','极大提升'], 'air_change': None},
         {'name': '旋转速度', 'dmg_b': 0, 'spd_b': 0.15, 'rng_b': 0, 'costs': [120,240,400,600,900],
          'tiers': ['加速转动','高速旋转','极速风车','飓风模式','永动风车'],
          'effects': ['攻速+15%','大幅提升','连续割击','不间断','攻速极限'], 'air_change': None},
         {'name': '风力范围', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [130,260,430,650,950],
          'tiers': ['扩展风场','大型风车','广域旋风','超级风场','全场飓风'],
          'effects': ['范围+15%','累计提升','大幅扩大','覆盖区域','无处遁形'], 'air_change': None},
         {'name': '气流效果', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [180,360,600,850,1200],
          'tiers': ['吹退气流','持续风压','强风锁定','吹飞轻型','龙卷风'],
          'effects': ['击退0.5s','移速-20%','眩晕0.5s','大幅击退','x2伤害+40%减速'], 'air_change': None},
     ]},
    {'name': '农夫', 'emoji': '👨‍🌾', 'rarity': '紫', 'cost': 500, 'dmg': 30, 'interval': 0.6, 'range': 280, 'atk_type': '全部', 'hero': '否', 'path_only': '否', 'unlock': 5,
     'paths': [
         {'name': '散弹枪', 'dmg_b': 0.15, 'spd_b': 0, 'rng_b': 0, 'costs': [200,400,650,900,1300],
          'tiers': ['加强装弹','穿透弹丸','爆炸弹丸','超级散射','霰弹之王'],
          'effects': ['弹丸+1','穿透1敌','小范围爆炸','弹丸翻倍+射程20%','大幅伤害+击退'], 'air_change': None},
         {'name': '机关枪', 'dmg_b': 0.10, 'spd_b': 0.15, 'rng_b': 0, 'costs': [180,360,600,850,1200],
          'tiers': ['半自动','全自动','重机枪','加特林','终极连射'],
          'effects': ['攻速+30%','攻速大幅提升','大型额外伤害','极高射速','攻速+伤害极限'], 'air_change': None},
         {'name': '狙击枪', 'dmg_b': 0.12, 'spd_b': 0, 'rng_b': 0.08, 'costs': [220,440,700,950,1400],
          'tiers': ['瞄准镜','穿透射击','爆头','反器材枪','神枪手'],
          'effects': ['射程+40%','穿透全部','25%双倍伤害','大型额外伤害','必暴击'], 'air_change': None},
         {'name': '爆头', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [350,600,900,1200,1600],
          'tiers': ['精准瞄准','要害射击','致命爆头','神经打击','终极狙杀'],
          'effects': ['20%暴击x2','30%暴击','40%+眩晕0.3s','50%+眩晕0.5s','60%+1s眩晕'], 'air_change': None},
     ]},
    {'name': '向日葵', 'emoji': '🌻', 'rarity': '紫', 'cost': 650, 'dmg': 8, 'interval': 1.5, 'range': 350, 'atk_type': '全部', 'hero': '否', 'path_only': '否', 'unlock': 60,
     'paths': [
         {'name': '阳光照耀', 'dmg_b': 0.15, 'spd_b': 0, 'rng_b': 0, 'costs': [350,700,1100,1500,2200],
          'tiers': ['晨光增幅','午后强光','烈日灼烧','黄金之光','神圣太阳'],
          'effects': ['伤害+15%','标记+5%','燃烧3s','标记增强','全场极限'], 'air_change': None},
         {'name': '光合速度', 'dmg_b': 0, 'spd_b': 0.10, 'rng_b': 0, 'costs': [300,600,950,1350,1950],
          'tiers': ['快速光合','高效日照','强化光能','极速日光','永恒太阳'],
          'effects': ['攻速+10%','累计提升','大幅提升','连续照射','攻速极限'], 'air_change': None},
         {'name': '照耀范围', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [320,640,1000,1400,2050],
          'tiers': ['延伸光芒','广域日照','大型太阳','超级辐射','星光普照'],
          'effects': ['射程+15%','累计提升','极大扩大','进一步提升','覆盖全图'], 'air_change': None},
         {'name': '黄金光线', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [400,800,1250,1750,2500],
          'tiers': ['光芒标记','增益光环','阳光领域','太阳之力','神圣庇护'],
          'effects': ['标记+20%/4s','标记30%+队友攻速10%','队友攻速15%','标记40%+攻速20%','全场标记+极限增益'], 'air_change': None},
     ]},
    {'name': '辣椒喷火器', 'emoji': '🌶️', 'rarity': '紫', 'cost': 480, 'dmg': 32, 'interval': 1.0, 'range': 220, 'atk_type': '地面', 'hero': '否', 'path_only': '否', 'unlock': 40,
     'paths': [
         {'name': '火焰强化', 'dmg_b': 0.18, 'spd_b': 0, 'rng_b': 0, 'costs': [200,400,640,880,1280],
          'tiers': ['浓缩辣素','高纯度辣油','极辣火焰','地狱辣椒','太阳炽热'],
          'effects': ['伤害+18%','燃烧+1s','燃烧伤害翻倍','大型双倍','伤害+燃烧极限'], 'air_change': None},
         {'name': '喷射速度', 'dmg_b': 0, 'spd_b': 0.15, 'rng_b': 0, 'costs': [160,320,520,720,1040],
          'tiers': ['增压泵','高效喷射','极速喷射','自动喷火','永恒火焰'],
          'effects': ['攻速+15%','累计提升','几乎连续','大幅提升','持续不断'], 'air_change': None},
         {'name': '喷射范围', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.12, 'costs': [180,360,580,800,1160],
          'tiers': ['延伸喷管','锥形扩展','宽幅火焰','广域喷射','烈焰洪流'],
          'effects': ['射程+12%','直线→锥形','多个并排','大幅提升','大范围锥形'], 'air_change': None},
         {'name': '辣椒标记', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [220,440,720,1000,1440],
          'tiers': ['辣味标记','火辣爆发','灼烧标记','终极辣椒','地狱魔辣'],
          'effects': ['标记+15%','标记20%+燃烧+2s','燃烧中+25%','减速20%','全炮台+40%'], 'air_change': None},
     ]},
    {'name': '弓弩', 'emoji': '🗼', 'rarity': '橙', 'cost': 600, 'dmg': 48, 'interval': 0.35, 'range': 500, 'atk_type': '地面', 'hero': '否', 'path_only': '否', 'unlock': 50,
     'paths': [
         {'name': '狙击威力', 'dmg_b': 0.15, 'spd_b': 0, 'rng_b': 0, 'costs': [300,600,1000,1400,2000],
          'tiers': ['精准射击','暴击弹头','穿甲狙击','必定暴击','终极狙击'],
          'effects': ['伤害+15%','25%暴击x2','无视25%护甲','必定暴击','穿透大型'], 'air_change': None},
         {'name': '射速提升', 'dmg_b': 0, 'spd_b': 0.10, 'rng_b': 0, 'costs': [250,500,800,1100,1600],
          'tiers': ['快速装弹','高效弹夹','连发模式','自动装弹','极速射击'],
          'effects': ['攻速+10%','减少换弹','连射2发','大幅提升','弹如雨点'], 'air_change': None},
         {'name': '超远射程', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.20, 'costs': [280,560,900,1250,1800],
          'tiers': ['瞄准镜','长管狙击','穿透射击','全图狙击','超级射程'],
          'effects': ['射程+20%','大幅提升','穿透1敌','穿透全部+全图','100%全图'], 'air_change': None},
         {'name': '标记猎杀', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [350,700,1100,1500,2200],
          'tiers': ['猎人标记','持久标记','暴击追踪','连锁标记','永久猎杀'],
          'effects': ['标记+20%/5s','持续10s','标记暴击+50%','击杀连锁标记','多目标+暴击+增伤'], 'air_change': None},
     ]},
    {'name': '老阿福', 'emoji': '🧑‍🌾', 'rarity': '橙', 'cost': 1200, 'dmg': 65, 'interval': 0.6, 'range': 300, 'atk_type': '全部', 'hero': '是', 'path_only': '否', 'unlock': 75,
     'paths': [
         {'name': '英雄力量', 'dmg_b': 0.20, 'spd_b': 0, 'rng_b': 0, 'costs': [350,700,1100,1500,2200],
          'tiers': ['农具强化','锻造大师','传奇武器','神器打造','农神之力'],
          'effects': ['伤害+20%','大型+30%','传奇品质','必附燃烧','必暴击+全效'], 'air_change': None},
         {'name': '英雄速度', 'dmg_b': 0, 'spd_b': 0.12, 'rng_b': 0, 'costs': [320,640,1000,1400,2050],
          'tiers': ['敏捷老农','双手操作','三连攻击','超速战斗','传说战士'],
          'effects': ['攻速+12%','动作迅猛','连击3下','额外多打','连击翻倍'], 'air_change': None},
         {'name': '英雄气场', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0.15, 'costs': [350,700,1100,1500,2200],
          'tiers': ['农场威严','领地扩张','守护者气场','英雄领域','农神庇护'],
          'effects': ['射程+15%','队友攻速+5%','队友+10%','队友+20%','全队极限增益'], 'air_change': None},
         {'name': '农场守护', 'dmg_b': 0, 'spd_b': 0, 'rng_b': 0, 'costs': [400,800,1250,1750,2500],
          'tiers': ['震慑一击','农夫之魂','田园守卫','传说英雄','农场之神'],
          'effects': ['20%眩晕1s','出血3dps+眩晕30%','出血+眩晕+燃烧','持续翻倍','大型100%+极限'], 'air_change': None},
     ]},
]

# ═══════════════════════════════════════════════════════════
# Sheet 1: 炮台总览
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = '炮台总览'
headers1 = ['炮台', 'Emoji', '品质', '底价', '基础伤害', '攻速(s)', '射程', '攻击类型', '基础DPS', '英雄', '路径放置', '解锁等级']
for c, h in enumerate(headers1, 1):
    ws1.cell(row=1, column=c, value=h)
style_header(ws1, 1, len(headers1))

for i, t in enumerate(TOWERS, 2):
    base_dps = round(t['dmg'] / t['interval'], 1) if t['interval'] > 0 else 0
    row_data = [t['name'], t['emoji'], t['rarity'], t['cost'], t['dmg'], t['interval'], t['range'], t['atk_type'], base_dps, t['hero'], t['path_only'], t['unlock']]
    for c, v in enumerate(row_data, 1):
        ws1.cell(row=i, column=c, value=v)
    style_data(ws1, i, len(headers1), RARITY_FILLS.get(t['rarity']))

# Also add Farm Guardian manually
guardian_row = len(TOWERS) + 2
gd = ['农场守卫者', '🗿', '橙', 550, 65, 1.2, 250, '地面', 54.2, '是', '否', 10]
for c, v in enumerate(gd, 1):
    ws1.cell(row=guardian_row, column=c, value=v)
style_data(ws1, guardian_row, len(headers1), RARITY_FILLS['橙'])

for c in range(1, len(headers1) + 1):
    ws1.column_dimensions[get_column_letter(c)].width = 14
ws1.column_dimensions['A'].width = 16

# ═══════════════════════════════════════════════════════════
# Sheet 2: 升级路线详情 (with per-tier DPS)
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet('升级路线详情')
headers2 = ['炮台', '路线', '级数', '升级名称', '效果描述', '本级费用', '累计费用(含底价)',
            '伤害加成%', '攻速加成%', '射程加成%', '有效伤害', '有效攻速(s)', 'DPS', '打空中']
for c, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=c, value=h)
style_header(ws2, 1, len(headers2))

row = 2
for t in TOWERS:
    base_air = '✅' if t['atk_type'] == '全部' else '❌'
    for p in t['paths']:
        cumulative_cost = t['cost']
        for tier in range(5):
            cumulative_cost += p['costs'][tier]
            tier_num = tier + 1
            dmg_pct = round(p['dmg_b'] * tier_num * 100, 1)
            spd_pct = round(p['spd_b'] * tier_num * 100, 1)
            rng_pct = round(p['rng_b'] * tier_num * 100, 1)
            eff_dmg = round(t['dmg'] * (1 + p['dmg_b'] * tier_num), 1)
            eff_interval = round(t['interval'] / (1 + p['spd_b'] * tier_num), 3) if (1 + p['spd_b'] * tier_num) > 0 else t['interval']
            dps = round(eff_dmg / eff_interval, 1) if eff_interval > 0 else 0
            # Air status
            air = base_air
            if p.get('air_change') and tier_num >= p['air_change']:
                air = '✅'
            row_data = [
                t['name'], p['name'], tier_num, p['tiers'][tier], p['effects'][tier],
                p['costs'][tier], cumulative_cost,
                f"+{dmg_pct}%" if dmg_pct > 0 else "-",
                f"+{spd_pct}%" if spd_pct > 0 else "-",
                f"+{rng_pct}%" if rng_pct > 0 else "-",
                eff_dmg, eff_interval, dps, air
            ]
            for c, v in enumerate(row_data, 1):
                ws2.cell(row=row, column=c, value=v)
            style_data(ws2, row, len(headers2), RARITY_FILLS.get(t['rarity']))
            row += 1
        # Empty separator row after each path
        row += 1

# Column widths
col_widths2 = [14, 14, 6, 14, 28, 10, 14, 10, 10, 10, 10, 10, 8, 8]
for c, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════════════════════════
# Sheet 3: DPS 排行榜
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet('DPS排行榜')
headers3 = ['排名', '炮台', '路线', '满级DPS', '总投入', 'DPS/千金', '打空中', '特殊效果']
for c, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=c, value=h)
style_header(ws3, 1, len(headers3))

# Collect all max-tier DPS entries
dps_entries = []
for t in TOWERS:
    base_air = '✅' if t['atk_type'] == '全部' else '❌'
    for p in t['paths']:
        total_cost = t['cost'] + sum(p['costs'])
        eff_dmg = t['dmg'] * (1 + p['dmg_b'] * 5)
        eff_interval = t['interval'] / (1 + p['spd_b'] * 5) if (1 + p['spd_b'] * 5) > 0 else t['interval']
        dps = round(eff_dmg / eff_interval, 1) if eff_interval > 0 else 0
        air = base_air
        if p.get('air_change'):
            air = '✅'
        special = p['effects'][4]
        dps_entries.append({
            'tower': t['name'], 'path': p['name'] + ' T5', 'dps': dps,
            'total_cost': total_cost, 'air': air, 'special': special,
            'rarity': t['rarity']
        })

# Add Farm Guardian
dps_entries.append({'tower': '农场守卫者', 'path': 'Lv10(英雄)', 'dps': 54.2, 'total_cost': 550, 'air': '✅(Lv3)', 'special': '溅射+冲击波+石头陷阱', 'rarity': '橙'})

dps_entries.sort(key=lambda x: x['dps'], reverse=True)

for i, e in enumerate(dps_entries, 2):
    rank = i - 1
    dps_per_k = round(e['dps'] / (e['total_cost'] / 1000), 1) if e['total_cost'] > 0 else 0
    row_data = [rank, e['tower'], e['path'], e['dps'], e['total_cost'], dps_per_k, e['air'], e['special']]
    for c, v in enumerate(row_data, 1):
        ws3.cell(row=i, column=c, value=v)
    style_data(ws3, i, len(headers3), RARITY_FILLS.get(e['rarity']))

col_widths3 = [6, 14, 18, 10, 10, 10, 8, 30]
for c, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════════════════════════
# Sheet 4: 敌人数据
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet('敌人数据')
headers4 = ['敌人', 'ID', 'HP', '速度', '金币掉落', '类型', '特殊能力']
for c, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=c, value=h)
style_header(ws4, 1, len(headers4))

enemies = [
    ['蚁群', 'ant_swarm', 25, 150, 2, '地面', '蚁后召唤'],
    ['小老鼠', 'rat_small', 60, 120, 5, '地面', '无'],
    ['蝗虫', 'locust', 40, 160, 3, '空中', '飞行'],
    ['松鼠', 'squirrel', 80, 140, 6, '地面', '无'],
    ['乌鸦', 'crow', 50, 130, 7, '空中', '飞行'],
    ['蛇', 'snake', 90, 100, 8, '地面', '无'],
    ['鼹鼠', 'mole', 120, 90, 10, '地面', '无视陷阱'],
    ['胖老鼠', 'rat_fat', 200, 80, 8, '地面', '无'],
    ['巨兔', 'giant_rabbit', 280, 100, 12, '地面', '无'],
    ['野猪', 'boar', 350, 110, 15, '地面', '冲锋2.5x速度3s'],
    ['蚁后', 'ant_queen', 250, 120, 20, '地面', '每1s召唤2蚁'],
    ['狐狸头领', 'fox_leader', 300, 190, 22, '地面', '10%闪避'],
    ['乌鸦王', 'crow_king', 400, 130, 25, '空中', '每10s召唤2乌鸦'],
    ['巨型鼹鼠', 'giant_mole', 600, 70, 28, '地面', '无视陷阱'],
    ['装甲野猪', 'armored_boar', 700, 75, 30, '地面', '冲锋2.5x/3s'],
    ['森林之王', 'forest_king', 3000, 100, 100, '空中', 'BOSS,25%HP护盾+召唤野猪'],
]

for i, e in enumerate(enemies, 2):
    for c, v in enumerate(e, 1):
        ws4.cell(row=i, column=c, value=v)
    style_data(ws4, i, len(headers4))

col_widths4 = [12, 16, 8, 8, 10, 8, 30]
for c, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(c)].width = w

# ═══════════════════════════════════════════════════════════
# Sheet 5: 40波数据
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet('40波数据')
headers5 = ['波次', '阶段', '波次名称', '敌人组成', '总数量', '预估金币', '首次出现']
for c, h in enumerate(headers5, 1):
    ws5.cell(row=1, column=c, value=h)
style_header(ws5, 1, len(headers5))

GOLD_MAP = {'rat_small':5,'rat_fat':8,'squirrel':6,'ant_swarm':2,'mole':10,'snake':8,
            'giant_rabbit':12,'boar':15,'armored_boar':30,'giant_mole':28,'fox_leader':22,
            'ant_queen':20,'crow':7,'locust':3,'crow_king':25,'forest_king':100}
NAME_MAP = {'rat_small':'小老鼠','rat_fat':'胖老鼠','squirrel':'松鼠','ant_swarm':'蚁群',
            'mole':'鼹鼠','snake':'蛇','giant_rabbit':'巨兔','boar':'野猪',
            'armored_boar':'装甲野猪','giant_mole':'巨型鼹鼠','fox_leader':'狐狸头领',
            'ant_queen':'蚁后','crow':'乌鸦','locust':'蝗虫','crow_king':'乌鸦王',
            'forest_king':'森林之王'}

WAVE_DATA = [
    [{'t':'rat_small','c':12}],
    [{'t':'rat_small','c':10},{'t':'rat_fat','c':5}],
    [{'t':'rat_small','c':8},{'t':'rat_fat','c':8},{'t':'squirrel','c':4}],
    [{'t':'rat_fat','c':10},{'t':'squirrel','c':6},{'t':'ant_swarm','c':10}],
    [{'t':'rat_fat','c':8},{'t':'squirrel','c':6},{'t':'ant_swarm','c':8},{'t':'mole','c':3}],
    [{'t':'ant_swarm','c':10},{'t':'mole','c':6},{'t':'snake','c':4},{'t':'giant_rabbit','c':2}],
    [{'t':'boar','c':3},{'t':'giant_rabbit','c':4},{'t':'snake','c':5},{'t':'ant_queen','c':2}],
    [{'t':'giant_rabbit','c':4},{'t':'snake','c':5},{'t':'mole','c':4}],
    [{'t':'armored_boar','c':4},{'t':'giant_mole','c':4},{'t':'fox_leader','c':3},{'t':'ant_queen','c':2}],
    [{'t':'armored_boar','c':5},{'t':'giant_mole','c':5},{'t':'fox_leader','c':4},{'t':'ant_queen','c':2}],
    [{'t':'snake','c':8},{'t':'mole','c':6},{'t':'crow','c':8}],
    [{'t':'locust','c':15},{'t':'giant_rabbit','c':6},{'t':'crow','c':9}],
    [{'t':'crow','c':10},{'t':'locust','c':12},{'t':'armored_boar','c':3},{'t':'snake','c':5}],
    [{'t':'giant_mole','c':4},{'t':'crow','c':10},{'t':'locust','c':12},{'t':'armored_boar','c':1}],
    [{'t':'ant_queen','c':2},{'t':'armored_boar','c':4},{'t':'crow','c':12},{'t':'locust','c':8}],
    [{'t':'giant_mole','c':5},{'t':'fox_leader','c':4},{'t':'crow','c':6},{'t':'locust','c':8}],
    [{'t':'armored_boar','c':5},{'t':'crow','c':12},{'t':'locust','c':15},{'t':'ant_queen','c':2}],
    [{'t':'giant_mole','c':5},{'t':'fox_leader','c':4},{'t':'armored_boar','c':2},{'t':'crow','c':8}],
    [{'t':'fox_leader','c':5},{'t':'armored_boar','c':4},{'t':'giant_mole','c':3},{'t':'crow','c':8}],
    [{'t':'crow_king','c':1},{'t':'armored_boar','c':5},{'t':'giant_mole','c':5},{'t':'crow','c':8}],
    [{'t':'ant_queen','c':4},{'t':'giant_mole','c':5},{'t':'fox_leader','c':4},{'t':'crow','c':8}],
    [{'t':'armored_boar','c':6},{'t':'giant_mole','c':5},{'t':'crow_king','c':1},{'t':'crow','c':8}],
    [{'t':'fox_leader','c':7},{'t':'armored_boar','c':4},{'t':'giant_mole','c':3},{'t':'locust','c':12}],
    [{'t':'crow_king','c':2},{'t':'crow','c':15},{'t':'locust','c':15},{'t':'armored_boar','c':5},{'t':'ant_queen','c':2}],
    [{'t':'armored_boar','c':6},{'t':'giant_mole','c':6},{'t':'fox_leader','c':4},{'t':'crow_king','c':1}],
    [{'t':'ant_queen','c':5},{'t':'giant_mole','c':6},{'t':'armored_boar','c':5},{'t':'fox_leader','c':2}],
    [{'t':'crow_king','c':3},{'t':'crow','c':15},{'t':'locust','c':20},{'t':'giant_mole','c':6}],
    [{'t':'armored_boar','c':6},{'t':'fox_leader','c':6},{'t':'giant_mole','c':4},{'t':'crow','c':6}],
    [{'t':'crow_king','c':2},{'t':'armored_boar','c':6},{'t':'giant_mole','c':6},{'t':'ant_queen','c':3}],
    [{'t':'crow_king','c':3},{'t':'armored_boar','c':5},{'t':'giant_mole','c':6},{'t':'fox_leader','c':4}],
    [{'t':'armored_boar','c':8},{'t':'fox_leader','c':6},{'t':'giant_mole','c':4},{'t':'crow_king','c':1}],
    [{'t':'ant_queen','c':5},{'t':'armored_boar','c':8},{'t':'giant_mole','c':5},{'t':'crow','c':6}],
    [{'t':'crow_king','c':4},{'t':'crow','c':20},{'t':'locust','c':25},{'t':'armored_boar','c':5},{'t':'giant_mole','c':3}],
    [{'t':'armored_boar','c':8},{'t':'giant_mole','c':8},{'t':'fox_leader','c':4},{'t':'crow_king','c':1}],
    [{'t':'armored_boar','c':7},{'t':'giant_mole','c':6},{'t':'fox_leader','c':6},{'t':'ant_queen','c':2}],
    [{'t':'crow_king','c':4},{'t':'armored_boar','c':6},{'t':'giant_mole','c':6},{'t':'fox_leader','c':4}],
    [{'t':'ant_queen','c':6},{'t':'armored_boar','c':7},{'t':'giant_mole','c':7},{'t':'crow_king','c':2}],
    [{'t':'crow_king','c':5},{'t':'crow','c':20},{'t':'armored_boar','c':8},{'t':'fox_leader','c':4}],
    [{'t':'armored_boar','c':8},{'t':'giant_mole','c':6},{'t':'fox_leader','c':6},{'t':'crow_king','c':2}],
    [{'t':'forest_king','c':1},{'t':'giant_mole','c':6},{'t':'armored_boar','c':5},{'t':'fox_leader','c':5},{'t':'ant_queen','c':3},{'t':'crow_king','c':1}],
]

WAVE_NAMES = [
    '热身：小鼠','肥鼠登场','松鼠加入','蚁群压阵','地鼠登场',
    '地鼠+大兔子','野猪首次登场','大兔子+地鼠','狐狸首领+蚁后','精英合围',
    '初见乌鸦','蝗虫过境','空中编队','双线夹击','蚁后空袭',
    '精英联队','铁翼冲锋','群兽乱野','暗影突击','乌鸦之王驾临(中BOSS)',
    '蚁穴风暴','铁甲洪流','狐狸精英','空中大军','全面攻势(里程碑)',
    '毒巢侵袭','暗翼军团','铁血精英','黑夜突袭','乌鸦之王·强化(中BOSS)',
    '末日序幕','全军出击','空中死神','铁幕压阵','巅峰军团',
    '怒潮席卷','终极蚁群','暗夜飞翼','绝境守护','森林之王(最终BOSS)',
]

PHASES = ['Phase1 农场入侵']*10 + ['Phase2 空中入侵']*10 + ['Phase3 精英阶段']*10 + ['Phase4 最终围攻']*10
PHASE_FILLS = {
    'Phase1 农场入侵': PatternFill('solid', fgColor='E8F5E9'),
    'Phase2 空中入侵': PatternFill('solid', fgColor='E3F2FD'),
    'Phase3 精英阶段': PatternFill('solid', fgColor='FFF3E0'),
    'Phase4 最终围攻': PatternFill('solid', fgColor='FFEBEE'),
}

seen_types = set()
for i, wave in enumerate(WAVE_DATA):
    row = i + 2
    wave_num = i + 1
    total_count = sum(g['c'] for g in wave)
    total_gold = sum(g['c'] * GOLD_MAP.get(g['t'], 5) for g in wave)
    composition = ', '.join(f"{NAME_MAP.get(g['t'],g['t'])}x{g['c']}" for g in wave)
    new_types = [NAME_MAP.get(g['t'],g['t']) for g in wave if g['t'] not in seen_types]
    for g in wave:
        seen_types.add(g['t'])
    first_appear = ', '.join(new_types) if new_types else '-'

    row_data = [wave_num, PHASES[i], WAVE_NAMES[i], composition, total_count, total_gold, first_appear]
    for c, v in enumerate(row_data, 1):
        ws5.cell(row=row, column=c, value=v)
    style_data(ws5, row, len(headers5), PHASE_FILLS.get(PHASES[i]))

col_widths5 = [6, 18, 24, 50, 8, 10, 20]
for c, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(c)].width = w

# Save
output_path = r'C:\sohai\阿福守农场\docs\农场防御塔_完整数据手册.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
