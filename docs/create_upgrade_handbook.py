import os, re, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

UPGRADES_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'global_upgrades')
OUTPUT = os.path.join(os.path.dirname(__file__), '波次强化升级_完整数据手册.xlsx')

RARITY_MAP = {0: '白(普通)', 1: '蓝(稀有)', 2: '橙(史诗)'}
TYPE_MAP = {0: '单炮台强化', 1: '全局强化', 2: '羁绊', 3: '建造折扣'}
STAT_MAP = {0: '伤害', 1: '攻速', 2: '射程', 3: '费用'}
TOWER_CN = {
    'farmer': '农夫', 'farm_cannon': '农场大炮', 'beehive': '蜂巢',
    'chili_flamer': '辣椒喷火器', 'mushroom_bomb': '蘑菇炸弹',
    'seed_shooter': '播种机', 'windmill': '风车', 'scarecrow': '稻草人',
    'water_pipe': '水压管道', 'watchtower': '弓弩', 'sunflower': '向日葵',
    'barbed_wire': '带刺铁网', 'bear_trap': '捕兽夹', 'hero_farmer': '老阿福',
    'farm_guardian': '农场守卫者',
}

def parse_tres(path):
    with open(path, 'r', encoding='utf-8') as f:
        text = f.read()
    d = {}
    for key in ['upgrade_id', 'display_name', 'description', 'icon_emoji',
                'target_tower_id']:
        m = re.search(rf'{key}\s*=\s*"([^"]*)"', text)
        d[key] = m.group(1) if m else ''
    for key in ['rarity', 'upgrade_type', 'stat_type']:
        m = re.search(rf'{key}\s*=\s*(\d+)', text)
        d[key] = int(m.group(1)) if m else 0
    m = re.search(r'stat_bonus\s*=\s*([\d.]+)', text)
    d['stat_bonus'] = float(m.group(1)) if m else 0.0
    # arrays - handle both [x] and Array[String]([x]) formats
    for key in ['required_tower_ids', 'target_tower_ids']:
        m = re.search(rf'{key}\s*=\s*(?:Array\[String\]\()?\[([^\]]*)\]', text)
        if m and m.group(1).strip():
            d[key] = [s.strip().strip('"') for s in m.group(1).split(',') if s.strip().strip('"')]
        else:
            d[key] = []
    return d

def tower_cn(tid):
    return TOWER_CN.get(tid, tid)

def towers_cn(ids):
    return ' + '.join(tower_cn(t) for t in ids) if ids else ''

# Parse all files
upgrades = []
for f in sorted(glob.glob(os.path.join(UPGRADES_DIR, '*.tres'))):
    upgrades.append(parse_tres(f))

tower_stat = [u for u in upgrades if u['upgrade_type'] == 0]
global_stat = [u for u in upgrades if u['upgrade_type'] == 1]
cost_red = [u for u in upgrades if u['upgrade_type'] == 3]
synergy_2 = [u for u in upgrades if u['upgrade_type'] == 2 and len(u['required_tower_ids']) <= 1]
synergy_m = [u for u in upgrades if u['upgrade_type'] == 2 and len(u['required_tower_ids']) > 1]

# Styles
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
HEADER_FILL_1 = PatternFill('solid', fgColor='2E7D32')  # green
HEADER_FILL_2 = PatternFill('solid', fgColor='1565C0')  # blue
HEADER_FILL_3 = PatternFill('solid', fgColor='E65100')  # orange
HEADER_FILL_4 = PatternFill('solid', fgColor='6A1B9A')  # purple
HEADER_FILL_5 = PatternFill('solid', fgColor='C62828')  # red
HEADER_FILL_6 = PatternFill('solid', fgColor='37474F')  # dark
DATA_FONT = Font(name='Arial', size=10)
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
WHITE_FILL = PatternFill('solid', fgColor='E8F5E9')
BLUE_FILL = PatternFill('solid', fgColor='E3F2FD')
ORANGE_FILL = PatternFill('solid', fgColor='FFF3E0')
RARITY_FILLS = {0: WHITE_FILL, 1: BLUE_FILL, 2: ORANGE_FILL}

def style_header(ws, row, cols, fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = ALIGN_C
        cell.border = THIN_BORDER

def style_data(ws, row, cols, rarity=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.alignment = ALIGN_C if c != cols else ALIGN_L
        cell.border = THIN_BORDER
        if rarity is not None and rarity in RARITY_FILLS:
            cell.fill = RARITY_FILLS[rarity]

wb = Workbook()

# ── Sheet 1: 单炮台强化 ──
ws1 = wb.active
ws1.title = '单炮台强化'
headers1 = ['ID', '名称', '图标', '品质', '目标炮台', '属性', '加成', '说明']
ws1.append(headers1)
style_header(ws1, 1, len(headers1), HEADER_FILL_1)
for u in sorted(tower_stat, key=lambda x: (x['target_tower_id'], x['rarity'])):
    row = [
        u['upgrade_id'], u['display_name'], u['icon_emoji'],
        RARITY_MAP[u['rarity']], tower_cn(u['target_tower_id']),
        STAT_MAP[u['stat_type']], f"+{int(u['stat_bonus']*100)}%",
        u['description']
    ]
    ws1.append(row)
    style_data(ws1, ws1.max_row, len(headers1), u['rarity'])
ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 6
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 14
ws1.column_dimensions['F'].width = 8
ws1.column_dimensions['G'].width = 8
ws1.column_dimensions['H'].width = 35

# ── Sheet 2: 全局强化 ──
ws2 = wb.create_sheet('全局强化')
headers2 = ['ID', '名称', '图标', '品质', '属性', '加成', '说明']
ws2.append(headers2)
style_header(ws2, 1, len(headers2), HEADER_FILL_2)
for u in sorted(global_stat, key=lambda x: x['rarity']):
    row = [
        u['upgrade_id'], u['display_name'], u['icon_emoji'],
        RARITY_MAP[u['rarity']], STAT_MAP[u['stat_type']],
        f"+{int(u['stat_bonus']*100)}%", u['description']
    ]
    ws2.append(row)
    style_data(ws2, ws2.max_row, len(headers2), u['rarity'])
for c in ['A','B','C','D','E','F','G']:
    ws2.column_dimensions[c].width = [20,18,6,12,8,8,35]['ABCDEFG'.index(c)]

# ── Sheet 3: 建造折扣 ──
ws3 = wb.create_sheet('建造折扣')
headers3 = ['ID', '名称', '图标', '品质', '折扣', '说明']
ws3.append(headers3)
style_header(ws3, 1, len(headers3), HEADER_FILL_3)
for u in sorted(cost_red, key=lambda x: x['rarity']):
    row = [
        u['upgrade_id'], u['display_name'], u['icon_emoji'],
        RARITY_MAP[u['rarity']], f"-{int(u['stat_bonus']*100)}%",
        u['description']
    ]
    ws3.append(row)
    style_data(ws3, ws3.max_row, len(headers3), u['rarity'])
for c in ['A','B','C','D','E','F']:
    ws3.column_dimensions[c].width = [20,18,6,12,8,35]['ABCDEF'.index(c)]

# ── Sheet 4: 双塔羁绊 ──
ws4 = wb.create_sheet('双塔羁绊')
headers4 = ['ID', '名称', '图标', '品质', '前置炮台', '受益炮台', '属性', '加成', '说明']
ws4.append(headers4)
style_header(ws4, 1, len(headers4), HEADER_FILL_4)
for u in sorted(synergy_2, key=lambda x: x['upgrade_id']):
    req = towers_cn(u['required_tower_ids'])
    target = towers_cn(u['target_tower_ids']) if u['target_tower_ids'] else tower_cn(u['target_tower_id'])
    row = [
        u['upgrade_id'], u['display_name'], u['icon_emoji'],
        RARITY_MAP[u['rarity']], req, target,
        STAT_MAP[u['stat_type']], f"+{int(u['stat_bonus']*100)}%",
        u['description']
    ]
    ws4.append(row)
    style_data(ws4, ws4.max_row, len(headers4), u['rarity'])
for i, w in enumerate([22,16,6,12,18,18,8,8,40]):
    ws4.column_dimensions[chr(65+i)].width = w

# ── Sheet 5: 多塔羁绊 ──
ws5 = wb.create_sheet('多塔羁绊')
headers5 = ['ID', '名称', '图标', '品质', '前置炮台(全部需要)', '受益炮台', '属性', '加成', '说明']
ws5.append(headers5)
style_header(ws5, 1, len(headers5), HEADER_FILL_5)
for u in sorted(synergy_m, key=lambda x: x['upgrade_id']):
    req = towers_cn(u['required_tower_ids'])
    target = towers_cn(u['target_tower_ids']) if u['target_tower_ids'] else tower_cn(u['target_tower_id'])
    row = [
        u['upgrade_id'], u['display_name'], u['icon_emoji'],
        RARITY_MAP[u['rarity']], req, target,
        STAT_MAP[u['stat_type']], f"+{int(u['stat_bonus']*100)}%",
        u['description']
    ]
    ws5.append(row)
    style_data(ws5, ws5.max_row, len(headers5), u['rarity'])
for i, w in enumerate([22,16,6,12,30,30,8,8,45]):
    ws5.column_dimensions[chr(65+i)].width = w

# ── Sheet 6: 系统规则 ──
ws6 = wb.create_sheet('系统规则')
rules = [
    ['波次强化系统规则', ''],
    ['', ''],
    ['触发波次', '第 5, 10, 15, 20, 25, 30, 35 波'],
    ['每次选择', '3选1'],
    ['单局上限', '8个强化'],
    ['', ''],
    ['品质权重', ''],
    ['白色(普通)', '60%'],
    ['蓝色(稀有)', '30%'],
    ['橙色(史诗)', '10%'],
    ['', ''],
    ['刷新费用', ''],
    ['金币刷新', '100金币'],
    ['钻石刷新', '5钻石'],
    ['', ''],
    ['统计', ''],
    ['单炮台强化', f'{len(tower_stat)}个'],
    ['全局强化', f'{len(global_stat)}个'],
    ['建造折扣', f'{len(cost_red)}个'],
    ['双塔羁绊', f'{len(synergy_2)}个'],
    ['多塔羁绊', f'{len(synergy_m)}个'],
    ['总计', f'{len(upgrades)}个'],
    ['', ''],
    ['羁绊激活条件', '场上必须已放置全部"前置炮台"才生效'],
    ['可叠加', '同一炮台可同时享受多个不同强化'],
    ['不可重复', '同一ID的强化一局只能选一次'],
]
for row in rules:
    ws6.append(row)
ws6['A1'].font = Font(name='Arial', bold=True, size=14)
for r in range(3, 26):
    ws6.cell(r, 1).font = Font(name='Arial', bold=True, size=11)
    ws6.cell(r, 2).font = DATA_FONT
ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 40

wb.save(OUTPUT)
print(f'Excel saved: {OUTPUT}')
print(f'Total upgrades: {len(upgrades)}')
print(f'  Tower stat: {len(tower_stat)}')
print(f'  Global stat: {len(global_stat)}')
print(f'  Cost reduction: {len(cost_red)}')
print(f'  Synergy (2-tower): {len(synergy_2)}')
print(f'  Synergy (multi): {len(synergy_m)}')
